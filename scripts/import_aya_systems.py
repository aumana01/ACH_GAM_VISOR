#!/usr/bin/env python3
"""Relaciona el SHP de sistemas AyA con el Excel ACH y genera la capa pública."""

from __future__ import annotations

import argparse
import gzip
import json
import math
import re
import struct
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "map" / "data" / "sistemas.geojson.gz"
DEFAULT_METADATA = ROOT / "map" / "data" / "metadata.json"
VALID_CATEGORIES = {"I", "II", "III", "IV", "SIN DATOS"}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalized_key(value: Any) -> str:
    text = unicodedata.normalize("NFD", clean(value).upper())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^A-Z0-9]", "", text)


def normalize_code(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", clean(value).upper())


def public_number(value: Any, digits: int) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise ValueError(f"Valor numérico inválido: {value!r}") from error
    if not math.isfinite(number):
        raise ValueError(f"Valor numérico no finito: {value!r}")
    return round(number, digits)


def read_workbook(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    raw_headers = next(rows)
    headers = [normalized_key(value) for value in raw_headers]
    required = {
        "CODIGOSISTEMA",
        "REGIONOPERATIVA",
        "NOMBREDELSISTEMA",
        "FACTOROCUPACIONPERSONASSERV",
        "CONSUMOPORSERVICIOM3MES",
        "DOTACIONESTIMADALPD",
        "CATEGORIASEGUNICH",
    }
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError("Faltan columnas requeridas en el Excel: " + ", ".join(missing))

    records: dict[str, dict[str, Any]] = {}
    for row_number, values in enumerate(rows, start=2):
        row = dict(zip(headers, values))
        code = normalize_code(row.get("CODIGOSISTEMA"))
        if not code:
            continue
        if code in records:
            raise ValueError(f"Código duplicado en el Excel: {code} (fila {row_number})")
        category = clean(row.get("CATEGORIASEGUNICH")).upper()
        if category not in VALID_CATEGORIES:
            raise ValueError(f"Categoría ICH inválida para {code}: {category or 'vacía'}")
        records[code] = {
            "codigo": code,
            "region": clean(row.get("REGIONOPERATIVA")),
            "nombre": clean(row.get("NOMBREDELSISTEMA")),
            "ich": category,
            "dotacion_lpd": public_number(row.get("DOTACIONESTIMADALPD"), 1),
            "consumo_conexion_m3_mes": public_number(
                row.get("CONSUMOPORSERVICIOM3MES"), 2
            ),
            "factor_ocupacion": public_number(
                row.get("FACTOROCUPACIONPERSONASSERV"), 2
            ),
        }
    workbook.close()
    return records


def read_dbf(path: Path) -> list[dict[str, Any]]:
    data = path.read_bytes()
    record_count = struct.unpack_from("<I", data, 4)[0]
    header_length = struct.unpack_from("<H", data, 8)[0]
    record_length = struct.unpack_from("<H", data, 10)[0]
    fields: list[tuple[str, str, int, int]] = []
    position = 32
    while position < header_length and data[position] != 0x0D:
        descriptor = data[position : position + 32]
        name = descriptor[:11].split(b"\x00", 1)[0].decode("ascii")
        fields.append((name, chr(descriptor[11]), descriptor[16], descriptor[17]))
        position += 32

    records: list[dict[str, Any]] = []
    offset = header_length
    for record_number in range(1, record_count + 1):
        record = data[offset : offset + record_length]
        offset += record_length
        if not record or record[0:1] == b"*":
            raise ValueError(f"El DBF contiene un registro eliminado: {record_number}")
        cursor = 1
        row: dict[str, Any] = {}
        for name, field_type, length, decimals in fields:
            raw = record[cursor : cursor + length]
            cursor += length
            text = raw.decode("utf-8", errors="replace").strip()
            if not text:
                value: Any = None
            elif field_type in {"N", "F"}:
                value = float(text) if decimals else int(text)
            else:
                value = text
            row[name] = value
        records.append(row)
    return records


def signed_area(ring: list[list[float]]) -> float:
    return 0.5 * sum(
        first[0] * second[1] - second[0] * first[1]
        for first, second in zip(ring, ring[1:])
    )


def ring_bounds(ring: list[list[float]]) -> tuple[float, float, float, float]:
    xs = [point[0] for point in ring]
    ys = [point[1] for point in ring]
    return min(xs), min(ys), max(xs), max(ys)


def point_in_ring(point: list[float], ring: list[list[float]]) -> bool:
    x, y = point
    inside = False
    previous = ring[-1]
    for current in ring:
        x1, y1 = previous
        x2, y2 = current
        if (y1 > y) != (y2 > y):
            intersection = (x2 - x1) * (y - y1) / (y2 - y1) + x1
            if x < intersection:
                inside = not inside
        previous = current
    return inside


def normalize_ring(points: list[tuple[float, float]], precision: int = 6) -> list[list[float]]:
    ring: list[list[float]] = []
    for longitude, latitude in points:
        current = [round(longitude, precision), round(latitude, precision)]
        if not ring or current != ring[-1]:
            ring.append(current)
    if ring and ring[0] != ring[-1]:
        ring.append(ring[0].copy())
    if len(ring) < 4:
        return []
    return ring


def rings_to_geometry(rings: list[list[list[float]]]) -> dict[str, Any]:
    valid = [ring for ring in rings if len(ring) >= 4 and signed_area(ring) != 0]
    if not valid:
        raise ValueError("La geometría no contiene anillos válidos.")

    # En un SHP de polígonos, los anillos exteriores son horarios y los huecos
    # antihorarios. Se conserva esa topología al construir Polygon/MultiPolygon.
    exteriors = [ring for ring in valid if signed_area(ring) < 0]
    holes = [ring for ring in valid if signed_area(ring) > 0]
    if not exteriors:
        exteriors, holes = valid, []

    polygons = [[ring] for ring in exteriors]
    exterior_info = [
        (index, ring_bounds(ring), abs(signed_area(ring)), ring)
        for index, ring in enumerate(exteriors)
    ]
    for hole in holes:
        point = hole[0]
        candidates = []
        for index, bounds, area, exterior in exterior_info:
            west, south, east, north = bounds
            if west <= point[0] <= east and south <= point[1] <= north:
                if point_in_ring(point, exterior):
                    candidates.append((area, index))
        if candidates:
            _, polygon_index = min(candidates)
            polygons[polygon_index].append(hole)
        else:
            # Una orientación excepcional no debe descartar área: el anillo se
            # publica como un polígono independiente y queda contabilizado.
            polygons.append([hole])

    if len(polygons) == 1:
        return {"type": "Polygon", "coordinates": polygons[0]}
    return {"type": "MultiPolygon", "coordinates": polygons}


def iter_shapefile(path: Path) -> Iterator[tuple[int, dict[str, Any], int, int]]:
    data = path.read_bytes()
    if struct.unpack_from(">i", data, 0)[0] != 9994:
        raise ValueError("El archivo no tiene una cabecera SHP válida.")
    declared_type = struct.unpack_from("<i", data, 32)[0]
    if declared_type not in {5, 15, 25}:
        raise ValueError(f"Se esperaba un SHP poligonal; tipo recibido: {declared_type}")

    position = 100
    while position + 8 <= len(data):
        record_number, content_words = struct.unpack_from(">ii", data, position)
        start = position + 8
        end = start + content_words * 2
        if end > len(data):
            raise ValueError(f"Registro SHP incompleto: {record_number}")
        shape_type = struct.unpack_from("<i", data, start)[0]
        if shape_type not in {5, 15, 25}:
            raise ValueError(f"Tipo geométrico no admitido en registro {record_number}")
        part_count, point_count = struct.unpack_from("<ii", data, start + 36)
        parts_start = start + 44
        parts = list(struct.unpack_from(f"<{part_count}i", data, parts_start))
        points_start = parts_start + part_count * 4
        raw_points = [
            struct.unpack_from("<2d", data, points_start + index * 16)
            for index in range(point_count)
        ]
        parts.append(point_count)
        rings = [
            normalize_ring(raw_points[parts[index] : parts[index + 1]])
            for index in range(part_count)
        ]
        geometry = rings_to_geometry([ring for ring in rings if ring])
        yield record_number, geometry, part_count, point_count
        position = end


def update_metadata(
    path: Path,
    records: dict[str, dict[str, Any]],
    feature_count: int,
) -> None:
    metadata = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}
    categories = Counter(record["ich"] for record in records.values())
    regions = sorted({record["region"] for record in records.values()})
    metadata.pop("period", None)
    metadata.pop("deficit", None)
    metadata.pop("surplus", None)
    metadata["generatedAt"] = datetime.now(timezone.utc).isoformat()
    metadata["systems"] = len(records)
    metadata["regions"] = len(regions)
    metadata["categoryCounts"] = dict(sorted(categories.items()))
    feature_counts = metadata.setdefault("featureCounts", {})
    feature_counts["systems"] = feature_count
    path.write_text(
        json.dumps(metadata, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--shapefile", required=True, type=Path)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--metadata", type=Path, default=DEFAULT_METADATA)
    args = parser.parse_args()

    shapefile = args.shapefile.resolve()
    dbf = shapefile.with_suffix(".dbf")
    if not shapefile.exists() or not dbf.exists():
        raise FileNotFoundError("El SHP y su DBF deben estar disponibles en la misma carpeta.")

    water = read_workbook(args.workbook.resolve())
    dbf_rows = read_dbf(dbf)
    shape_codes = [normalize_code(row.get("codsistema")) for row in dbf_rows]
    if len(shape_codes) != len(set(shape_codes)):
        raise ValueError("El SHP contiene códigos de sistema duplicados.")
    missing_in_excel = sorted(set(shape_codes) - set(water))
    missing_in_shp = sorted(set(water) - set(shape_codes))
    if missing_in_excel or missing_in_shp:
        raise ValueError(
            "La relación por código no es completa. "
            f"Solo SHP: {missing_in_excel}; solo Excel: {missing_in_shp}"
        )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    feature_count = 0
    part_total = 0
    point_total = 0
    with gzip.open(args.output, "wt", encoding="utf-8", compresslevel=9) as target:
        target.write('{"type":"FeatureCollection","features":[')
        for index, (record_number, geometry, parts, points) in enumerate(
            iter_shapefile(shapefile)
        ):
            if index >= len(dbf_rows):
                raise ValueError("El SHP tiene más registros que el DBF.")
            code = shape_codes[index]
            properties = water[code]
            feature = {
                "type": "Feature",
                "geometry": geometry,
                "properties": properties,
            }
            if feature_count:
                target.write(",")
            target.write(
                json.dumps(feature, ensure_ascii=False, separators=(",", ":"))
            )
            feature_count += 1
            part_total += parts
            point_total += points
            if record_number != feature_count:
                raise ValueError("El orden de registros del SHP no es consecutivo.")
        target.write("]}")

    if feature_count != len(dbf_rows):
        raise ValueError("El SHP y el DBF no tienen la misma cantidad de registros.")
    update_metadata(args.metadata, water, feature_count)
    print(
        json.dumps(
            {
                "status": "ok",
                "systems": feature_count,
                "regions": len({record["region"] for record in water.values()}),
                "categories": dict(Counter(record["ich"] for record in water.values())),
                "parts": part_total,
                "sourcePoints": point_total,
                "outputBytes": args.output.stat().st_size,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
