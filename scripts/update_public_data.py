#!/usr/bin/env python3
"""Genera GeoJSON públicos depurados a partir de fuentes privadas actualizadas."""

from __future__ import annotations

import json
import os
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from shapely import voronoi_polygons
from shapely.geometry import MultiPoint, mapping, shape
from shapely.ops import unary_union


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = Path(os.environ.get("ACH_GAM_SOURCE_DIR", ROOT / "source_private"))
OUTPUT_DIR = ROOT / "map" / "data"
GAM_BOUNDS = (-84.66, 9.48, -83.76, 10.19)

FILES = {
    "systems": "Sistemas_y_Zonas_de_Abastecimiento.json",
    "water": "DATOS HÍDRICOS PUBLICO.xlsx",
    "municipal": "Acueductos_Municipales.json",
    "esph": "ESPH_AP.json",
    "asadas": "ASADAS.json",
    "thiessen": "Cobertura_Thiessen_ASADAS_UTAPS.json",
    "criteria": "Criterios Especiales CCH GAM.json",
    "ona": "Cobertura_ONAs_BD.json",
    "protected": "Áreas_Protegidas.json",
    "districts": "Distritos_GAM.json",
}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalized_text(value: Any) -> str:
    plain = unicodedata.normalize("NFD", clean(value))
    return "".join(char for char in plain if unicodedata.category(char) != "Mn")


def normalize_code(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", clean(value).upper())


def public_code(value: Any) -> str:
    normalized = normalize_code(value)
    match = re.fullmatch(r"MEA(\d{2})", normalized)
    return f"ME-A-{match.group(1)}" if match else clean(value)


def finite_number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number


def special_condition(value: Any) -> dict[str, str]:
    raw = clean(value)
    plain = normalized_text(raw)
    article = re.search(r"articulo\s*(\d+)", plain, re.IGNORECASE)
    if article:
        return {"tipo": "Facilidad", "detalle": f"Artículo {article.group(1)}"}
    if re.search(r"restric", plain, re.IGNORECASE):
        return {"tipo": "Restricción", "detalle": "Restricción particular"}
    return {"tipo": "Criterio especial", "detalle": raw or "No especificado"}


def read_geojson(filename: str) -> dict[str, Any]:
    path = SOURCE_DIR / filename
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if data.get("type") != "FeatureCollection" or not isinstance(
        data.get("features"), list
    ):
        raise ValueError(f"{filename} no es una FeatureCollection GeoJSON válida.")
    return data


def read_water_data() -> dict[str, dict[str, Any]]:
    workbook = load_workbook(SOURCE_DIR / FILES["water"], read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    result: dict[str, dict[str, Any]] = {}

    def value_with_prefix(row: dict[str, Any], prefix: str) -> Any:
        target = normalized_text(prefix).lower()
        for key, value in row.items():
            if normalized_text(key).lower().startswith(target):
                return value
        return None

    for values in rows:
        row = dict(zip(headers, values))
        code = normalize_code(row.get("Código del Sistema") or row.get("COD"))
        if not code:
            continue
        balance = finite_number(value_with_prefix(row, "balance del sistema"))
        ich = clean(value_with_prefix(row, "clasificacion ich"))
        period = finite_number(row.get("Período"))
        result[code] = {
            "code": public_code(code),
            "name": clean(row.get("Nombre del Sistema")),
            "condition": (
                "Sin clasificación"
                if balance is None
                else "Déficit" if balance < 0 else "Superávit"
            ),
            "ich": ich if ich in {"I", "II", "III", "IV"} else "Sin clasificación",
            "period": int(period) if period is not None else None,
        }
    workbook.close()
    return result


def coordinate_bounds(value: Any, bounds: list[float] | None = None) -> list[float]:
    current = bounds or [float("inf"), float("inf"), float("-inf"), float("-inf")]
    if not isinstance(value, list):
        return current
    if (
        len(value) >= 2
        and isinstance(value[0], (int, float))
        and isinstance(value[1], (int, float))
    ):
        current[0] = min(current[0], value[0])
        current[1] = min(current[1], value[1])
        current[2] = max(current[2], value[0])
        current[3] = max(current[3], value[1])
        return current
    for child in value:
        coordinate_bounds(child, current)
    return current


def intersects_gam(feature: dict[str, Any]) -> bool:
    coordinates = (feature.get("geometry") or {}).get("coordinates")
    if not coordinates:
        return False
    minimum_x, minimum_y, maximum_x, maximum_y = coordinate_bounds(coordinates)
    west, south, east, north = GAM_BOUNDS
    return not (
        maximum_x < west
        or minimum_x > east
        or maximum_y < south
        or minimum_y > north
    )


def round_coordinates(value: Any, precision: int) -> Any:
    if not isinstance(value, (list, tuple)):
        return value
    if (
        len(value) >= 2
        and isinstance(value[0], (int, float))
        and isinstance(value[1], (int, float))
    ):
        return [round(number, precision) if isinstance(number, float) else number for number in value]
    return [round_coordinates(child, precision) for child in value]


def map_features(
    collection: dict[str, Any],
    property_mapper: Callable[[dict[str, Any]], dict[str, Any]],
    *,
    within_gam: bool = False,
    tolerance: float = 0,
    precision: int = 6,
) -> dict[str, Any]:
    features = []
    for source_feature in collection["features"]:
        if not source_feature.get("geometry"):
            continue
        if within_gam and not intersects_gam(source_feature):
            continue
        geometry = source_feature["geometry"]
        if tolerance:
            simplified = shape(geometry).simplify(tolerance, preserve_topology=True)
            if simplified.is_empty:
                continue
            geometry = mapping(simplified)
        coordinates = round_coordinates(geometry.get("coordinates"), precision)
        features.append(
            {
                "type": "Feature",
                "geometry": {"type": geometry["type"], "coordinates": coordinates},
                "properties": property_mapper(source_feature.get("properties") or {}),
            }
        )
    return {"type": "FeatureCollection", "features": features}


def write_json(filename: str, data: dict[str, Any]) -> None:
    (OUTPUT_DIR / filename).write_text(
        json.dumps(data, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )


missing = [filename for filename in FILES.values() if not (SOURCE_DIR / filename).exists()]
if missing:
    raise FileNotFoundError(
        "Faltan archivos en la carpeta de fuentes: " + ", ".join(sorted(missing))
    )

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
water = read_water_data()

systems = map_features(
    read_geojson(FILES["systems"]),
    lambda source: {
        "codigo": water.get(normalize_code(source.get("Codigo_Sis")), {}).get("code")
        or public_code(source.get("Codigo_Sis")),
        "nombre": water.get(normalize_code(source.get("Codigo_Sis")), {}).get("name")
        or clean(source.get("Nombre_Sis")),
        "condicion": water.get(normalize_code(source.get("Codigo_Sis")), {}).get("condition")
        or ("Déficit" if (finite_number(source.get("Balance")) or 0) < 0 else "Superávit"),
        "ich": water.get(normalize_code(source.get("Codigo_Sis")), {}).get("ich")
        or clean(source.get("ICH")),
    },
    tolerance=0.000055,
)

municipal = map_features(
    read_geojson(FILES["municipal"]),
    lambda source: {
        "operador": clean(source.get("Operador")) or "Acueducto municipal",
        "sistema": clean(source.get("Sistema")) or "Sin nombre",
    },
    within_gam=True,
    tolerance=0.00022,
    precision=5,
)

esph = map_features(
    read_geojson(FILES["esph"]),
    lambda _source: {
        "operador": "Empresa de Servicios Públicos de Heredia (ESPH)",
        "sistema": "Cobertura de agua potable ESPH",
    },
    within_gam=True,
    tolerance=0.00012,
    precision=5,
)

asadas = map_features(
    read_geojson(FILES["asadas"]),
    lambda source: {
        "codigo": clean(source.get("CODIGO_IDEO") or source.get("IDEO")),
        "operador": clean(source.get("NOMBRE_DEL_OPERADOR") or source.get("Ente_Operador"))
        or "ASADA",
    },
    within_gam=True,
)

onas = map_features(
    read_geojson(FILES["ona"]),
    lambda source: {
        "operador": clean(source.get("Operador")) or "Operador local",
        "sistema": clean(source.get("Sistema")) or "Sin nombre",
    },
    within_gam=True,
    tolerance=0.00018,
    precision=5,
)

protected = map_features(
    read_geojson(FILES["protected"]),
    lambda source: {
        "codigo": clean(source.get("Codigo")),
        "nombre": clean(source.get("Nombre")) or "Área protegida",
        "categoria": clean(source.get("Categoria")),
    },
    within_gam=True,
    tolerance=0.0002,
    precision=5,
)

districts = map_features(
    read_geojson(FILES["districts"]),
    lambda source: {
        "provincia": clean(source.get("provincia")),
        "canton": clean(source.get("canton")),
        "distrito": clean(source.get("distrito")),
    },
    within_gam=True,
    tolerance=0.00012,
    precision=5,
)

systems_by_code: dict[str, dict[str, Any]] = {}
for feature in systems["features"]:
    code = normalize_code(feature["properties"].get("codigo"))
    if code and code not in systems_by_code:
        systems_by_code[code] = feature

criteria_features = []
for source_feature in read_geojson(FILES["criteria"])["features"]:
    source = source_feature.get("properties") or {}
    system = systems_by_code.get(normalize_code(source.get("codigo_sis")))
    if not system:
        continue
    condition = special_condition(source.get("cond_especial"))
    criteria_features.append(
        {
            "type": "Feature",
            "geometry": system["geometry"],
            "properties": {
                "codigo_sistema": system["properties"]["codigo"],
                "nombre_sistema": system["properties"]["nombre"],
                "codigo_abastecimiento": clean(source.get("codigo_aba")),
                "zona": clean(source.get("zonas")),
                "zona_operativa": clean(source.get("zona_opera")),
                **condition,
            },
        }
    )
criteria = {"type": "FeatureCollection", "features": criteria_features}

# La fuente Thiessen suministrada puede llegar sin geometría. Se valida su
# presencia y se reconstruye una cobertura estimada a partir de los puntos
# públicos de ASADAS, limitada a los distritos GAM.
thiessen_source = read_geojson(FILES["thiessen"])
if not thiessen_source["features"]:
    raise ValueError(f"{FILES['thiessen']} no contiene entidades.")

district_mask = unary_union(
    shape(feature["geometry"]) for feature in districts["features"]
).buffer(0)
unique_points: dict[tuple[float, float], dict[str, Any]] = {}
for feature in asadas["features"]:
    point = shape(feature["geometry"])
    if not district_mask.covers(point):
        continue
    key = (round(point.x, 7), round(point.y, 7))
    record = unique_points.setdefault(
        key,
        {"point": point, "codigos": set(), "operadores": set()},
    )
    properties = feature.get("properties") or {}
    if clean(properties.get("codigo")):
        record["codigos"].add(clean(properties.get("codigo")))
    if clean(properties.get("operador")):
        record["operadores"].add(clean(properties.get("operador")))

point_records = list(unique_points.values())
cells = voronoi_polygons(
    MultiPoint([record["point"] for record in point_records]),
    extend_to=district_mask,
    ordered=True,
)
thiessen_features = []
for record, cell in zip(point_records, cells.geoms):
    clipped = cell.intersection(district_mask).simplify(
        0.00008, preserve_topology=True
    )
    if clipped.is_empty:
        continue
    geometry = mapping(clipped)
    thiessen_features.append(
        {
            "type": "Feature",
            "geometry": {
                "type": geometry["type"],
                "coordinates": round_coordinates(geometry["coordinates"], 6),
            },
            "properties": {
                "codigo": " / ".join(sorted(record["codigos"])),
                "operador": " / ".join(sorted(record["operadores"])) or "ASADA",
                "alcance": "Cobertura somera/estimada",
                "metodo": "Polígono de Thiessen",
            },
        }
    )
thiessen = {"type": "FeatureCollection", "features": thiessen_features}

outputs = {
    "sistemas.geojson": systems,
    "municipalidades.geojson": municipal,
    "esph.geojson": esph,
    "asadas.geojson": asadas,
    "cobertura-thiessen-asadas.geojson": thiessen,
    "criterios-especiales.geojson": criteria,
    "onas.geojson": onas,
    "areas-protegidas.geojson": protected,
    "distritos.geojson": districts,
}
for filename, collection in outputs.items():
    write_json(filename, collection)

unique_systems = list(water.values())
periods = [item["period"] for item in unique_systems if item.get("period")]
metadata = {
    "generatedAt": datetime.now(timezone.utc).isoformat(),
    "period": max(periods) if periods else None,
    "systems": len(unique_systems),
    "deficit": sum(item["condition"] == "Déficit" for item in unique_systems),
    "surplus": sum(item["condition"] == "Superávit" for item in unique_systems),
    "featureCounts": {
        filename.removesuffix(".geojson"): len(collection["features"])
        for filename, collection in outputs.items()
    },
}
write_json("metadata.json", metadata)
print(json.dumps(metadata, ensure_ascii=False, indent=2))
